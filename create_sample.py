#!/usr/bin/env python3
"""
Creates a sample Excel file (sample_urls.xlsx) for testing Shorts Analyzer.
Run once: python create_sample.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "URLs"

# Header
ws["A1"] = "Video URL"
ws["A1"].font = Font(bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center")
ws.column_dimensions["A"].width = 60

# Sample URLs (replace with real ones)
sample_urls = [
    "https://www.youtube.com/shorts/REPLACE_WITH_REAL_ID",
    "https://www.tiktok.com/@username/video/REPLACE_WITH_REAL_ID",
    "https://www.instagram.com/reel/REPLACE_WITH_REAL_ID/",
]

for i, url in enumerate(sample_urls, start=2):
    ws.cell(row=i, column=1).value = url

wb.save("sample_urls.xlsx")
print("Created sample_urls.xlsx — replace the placeholder URLs with real ones.")
