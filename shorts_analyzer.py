#!/usr/bin/env python3
"""
Shorts Analyzer — Excel Batch Processor
엑셀 파일 내 모든 URL을 자동 감지하여 분석하고,
각 URL 열 바로 우측에 결과 열을 삽입한 새 파일을 출력합니다.
원본 파일의 내용과 디자인을 그대로 유지합니다.

Usage:
    python shorts_analyzer.py list_check.xlsx
    python shorts_analyzer.py list_check.xlsx --api http://localhost:8000
    python shorts_analyzer.py list_check.xlsx --session YOUR_SESSION_ID
"""

import sys
import re
import json
import time
import shutil
import argparse
import urllib.request
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────

DEFAULT_API    = "https://shorts-analyzer-api-production.up.railway.app"
METRIC_HEADERS = ["Views", "Likes", "Comments", "Shares", "Status"]
METRIC_WIDTHS  = [12, 10, 10, 10, 24]

HEADER_FILL  = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
OK_FILL      = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ERR_FILL     = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)

# ── API call ──────────────────────────────────────────────────────────────────

def analyze_urls(urls: list, api_base: str, session_id: str = "",
                 delay: float = 0.3, concurrent: int = 5) -> dict:
    """URL 목록을 API로 보내고 {url: result} 딕셔너리 반환."""
    if not urls:
        return {}
    body = json.dumps({
        "urls": urls,
        "delay": delay,
        "concurrent": concurrent,
        "yt_key": "",
        "insta_session": session_id,
    }).encode()
    req = urllib.request.Request(
        f"{api_base.rstrip('/')}/analyze",
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    results = {}
    with urllib.request.urlopen(req, timeout=600) as r:
        for raw in r:
            line = raw.decode("utf-8", "replace").strip()
            if not line.startswith("data:"):
                continue
            payload = line[5:].strip()
            if not payload:
                continue
            obj = json.loads(payload)
            if obj.get("done"):
                continue
            results[obj["url"]] = obj
    return results


def sanitize(val):
    """음수값 → None."""
    if val is None:
        return None
    if isinstance(val, (int, float)) and val < 0:
        return None
    return val


def extract_first_url(cell_text: str):
    """셀 텍스트에서 첫 번째 http URL 추출."""
    for part in re.split(r"[\s,、\n]+", cell_text):
        part = part.strip()
        if part.startswith("http"):
            return part
    return None


def platform_prefix(url: str) -> str:
    u = url.lower()
    if "tiktok.com" in u:                          return "TT_"
    if "youtube.com" in u or "youtu.be" in u:      return "YT_"
    if "instagram.com" in u:                       return "IG_"
    if "x.com" in u or "twitter.com" in u:         return "X_"
    if "lipscosme.com" in u:                       return "LIPS_"
    if "cosme.net" in u:                           return "AtC_"
    if "lemon8" in u:                              return "L8_"
    if "bsky.app" in u:                            return "Sky_"
    if "threads.com" in u or "threads.net" in u:   return "TH_"
    if "note.com" in u:                            return "Note_"
    return ""


def detect_col_prefix(col_urls: list) -> str:
    """
    URL 목록에서 지배적 플랫폼(≥80%)이 있으면 해당 접두어를 반환.
    혼합 플랫폼이거나 알 수 없는 플랫폼이면 빈 문자열 반환.
    """
    if not col_urls:
        return ""
    counts: dict = {}
    for url in col_urls:
        p = platform_prefix(url)
        counts[p] = counts.get(p, 0) + 1

    top_prefix, top_count = max(counts.items(), key=lambda x: x[1])
    total = len(col_urls)

    # 단일 플랫폼이거나 하나의 플랫폼이 80% 이상이고 접두어가 있으면 반환
    if len(counts) == 1 and top_prefix:
        return top_prefix
    if top_prefix and top_count / total >= 0.80:
        return top_prefix
    return ""


# ── Main processor ────────────────────────────────────────────────────────────

def process_excel(input_path: str, output_path: str = None,
                  api_base: str = DEFAULT_API, session_id: str = "",
                  delay: float = 0.3, concurrent: int = 5):

    p = Path(input_path)
    if output_path is None:
        output_path = str(p.parent / f"{p.stem}_result{p.suffix}")

    # ── 1단계: URL 읽기 ─────────────────────────────────────────────────────
    wb_r = openpyxl.load_workbook(input_path, data_only=True)
    ws_r = wb_r.active
    max_row = ws_r.max_row
    max_col = ws_r.max_column

    # 모든 열 스캔 → URL이 하나라도 있으면 URL 열로 등록
    url_cols = []
    for col in range(1, max_col + 1):
        for row in range(2, max_row + 1):          # 헤더(1행) 제외
            val = ws_r.cell(row=row, column=col).value
            if val and extract_first_url(str(val).strip()):
                url_cols.append(col)
                break

    print(f"원본: {max_row}행 × {max_col}열")
    print(f"URL 열 감지: {url_cols} ({len(url_cols)}개)")

    # 각 (row, col) → 대표 URL 매핑
    cell_url = {}   # (row, col) → url
    url_set  = set()
    for col in url_cols:
        for row in range(2, max_row + 1):
            val = ws_r.cell(row=row, column=col).value
            if not val:
                continue
            url = extract_first_url(str(val).strip())
            if url:
                cell_url[(row, col)] = url
                url_set.add(url)

    wb_r.close()
    unique_urls = list(url_set)
    print(f"분석 URL: {len(unique_urls)}개 (고유)\n")

    # ── 2단계: API 일괄 호출 ────────────────────────────────────────────────
    print(f"API 호출 중 ({api_base}) …")
    all_results = {}
    batch_size  = 50
    for i in range(0, len(unique_urls), batch_size):
        batch = unique_urls[i:i + batch_size]
        print(f"  배치 {i // batch_size + 1}: {len(batch)}개", flush=True)
        try:
            all_results.update(
                analyze_urls(batch, api_base, session_id, delay, concurrent)
            )
        except Exception as e:
            print(f"  ⚠ 배치 실패: {e}")
        if i + batch_size < len(unique_urls):
            time.sleep(1)

    ok = sum(1 for r in all_results.values() if r.get("status") == "Success")
    print(f"분석 완료: {ok}/{len(unique_urls)} 성공\n")

    # ── 3단계: 원본 파일 복사 후 열 삽입 ────────────────────────────────────
    # 원본을 그대로 복사해 모든 서식·병합·행높이 보존
    shutil.copy2(input_path, output_path)
    wb_w = openpyxl.load_workbook(output_path)
    ws_w = wb_w.active

    # 오른쪽 열부터 처리해야 왼쪽 열 인덱스가 밀리지 않음
    cumulative = 0   # 지금까지 삽입된 총 열 수

    for orig_col in sorted(url_cols):
        # 이미 삽입된 열들을 감안한 실제 현재 위치
        actual_col = orig_col + cumulative

        # URL 열의 모든 URL 수집 → 플랫폼 접두어 결정
        col_urls_list = [
            cell_url[(r, orig_col)]
            for r in range(2, max_row + 1)
            if (r, orig_col) in cell_url
        ]
        prefix = detect_col_prefix(col_urls_list)

        # 헤더 텍스트로 플랫폼 보완 (URL보다 헤더가 더 명확한 경우)
        col_header_val = ws_w.cell(row=1, column=actual_col).value or ""
        col_header_l   = col_header_val.lower() if col_header_val else ""
        if "tiktok"    in col_header_l: prefix = "TT_"
        elif "youtube" in col_header_l: prefix = "YT_"
        elif "instagram" in col_header_l: prefix = "IG_"
        elif "twitter" in col_header_l or "x(旧" in col_header_l or "x(" in col_header_l:
            prefix = "X_"

        # URL 열 바로 뒤에 메트릭 열 삽입
        insert_at = actual_col + 1
        n_cols    = len(METRIC_HEADERS)
        ws_w.insert_cols(insert_at, n_cols)

        # 헤더 행 작성
        for i, metric in enumerate(METRIC_HEADERS):
            c = ws_w.cell(row=1, column=insert_at + i)
            c.value     = f"{prefix}{metric}"
            c.font      = HEADER_FONT
            c.fill      = HEADER_FILL
            c.alignment = CENTER
            ws_w.column_dimensions[get_column_letter(insert_at + i)].width = METRIC_WIDTHS[i]

        # 데이터 행 작성
        for row in range(2, max_row + 1):
            url = cell_url.get((row, orig_col))
            if not url:
                continue
            result = all_results.get(url)
            if result is None:
                # API 결과 없음 — 빈 셀
                continue
            is_ok = result.get("status") == "Success"
            fill  = OK_FILL if is_ok else ERR_FILL
            vals  = [
                sanitize(result.get("views")),
                sanitize(result.get("likes")),
                sanitize(result.get("comments")),
                sanitize(result.get("shares")),
                result.get("status", "Error"),
            ]
            for i, v in enumerate(vals):
                c = ws_w.cell(row=row, column=insert_at + i)
                c.value     = v
                c.fill      = fill
                c.alignment = CENTER

        cumulative += n_cols

    wb_w.save(output_path)
    total_cols = max_col + cumulative
    print(f"저장 완료: {output_path}")
    print(f"최종: {max_row}행 × {total_cols}열 (메트릭 {cumulative}열 추가)")
    return output_path


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Shorts Analyzer — 엑셀 파일의 모든 URL을 분석해 결과 열 삽입"
    )
    parser.add_argument("input",              help="입력 엑셀 파일 (.xlsx)")
    parser.add_argument("-o", "--output",     default=None, help="출력 파일 경로")
    parser.add_argument("--api",              default=DEFAULT_API, help=f"API 주소 (기본: {DEFAULT_API})")
    parser.add_argument("--session",          default="",   help="Instagram Session ID")
    parser.add_argument("--delay",            type=float, default=0.3, help="요청 간 딜레이 (초)")
    parser.add_argument("--concurrent",       type=int,   default=5,   help="동시 요청 수")

    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"Error: 파일 없음 — {args.input}")
        sys.exit(1)

    process_excel(
        input_path   = args.input,
        output_path  = args.output,
        api_base     = args.api,
        session_id   = args.session,
        delay        = args.delay,
        concurrent   = args.concurrent,
    )


if __name__ == "__main__":
    main()
